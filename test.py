# convert A-Z AA-AZ BA-BZ ... to 1-26 27-52 53-78 ...
# def convert_excel_column_to_number(column: str):
#     column = column.upper()
#     column_number = 0
#     for i in range(len(column)):
#         column_number += (ord(column[i]) - 64) * (26 ** (len(column) - i - 1))
#     return column_number
#
# print(convert_excel_column_to_number("CP"))

import openpyxl
# sheet_name = "PM值 先轉成數字 再貼"
sheet_name = "清單"
workbook = openpyxl.load_workbook("data/BI.xlsx")
worksheet = workbook[sheet_name]

# get max row and column
max_row = worksheet.max_row
max_column = worksheet.max_column

# setting value
for row in range(1, max_row + 1):
    print("-" * 100)
    for column in range(1, max_column + 1):
        print(f"{column}. " + "*" * 20)
        # if column == max_column:
        #     cell = worksheet.cell(row, column + 1)
        #     # set value as VBA formula 『=IF(CP2<-10,CP2,IF(ISERROR(AVERAGE(CN2:CR2)),"",AVERAGE(CN2:CR2)))』
        #     cell.value = f'=IF(CP2<-10,CP2,IF(ISERROR(AVERAGE(CN2:CR2)),"",AVERAGE(CN2:CR2)))'
        print(worksheet.cell(row, column).value)
    if row == 2:
        break