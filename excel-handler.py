import os
import time
import shutil
import openpyxl
from tqdm import tqdm
from configparser import ConfigParser


from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    borders,
)

COL_NAMES = [
    None,
    "MDT RSRP",
    "site1昨日最差PRB",
    "site2昨日最差PRB",
    "site3昨日最差PRB",
    "昨日最差PRB",
    "昨日最差PRB的SiteID	",
    "Weekday",
    "RSRP(MDT)",
    None,
    "5G True User",
    "Site1~3 近3小時干擾>-105的筆數",
    "昨日最差PRB(4)",
    "RSRP(4)",
    "新客訴原因4",
    "RuleBase",
    "OM回覆客訴原因",
]


def main(input_file: str, sheet_name: str, output_file: str, output_sheet_name: str):
    # check if output.xlsx exists
    isExists = False
    if os.path.isfile(f"{output_file}"):  # 檢查BI檔是否存在，如果有把da的資料加到BI裡面
        isExists = True
        original_max_rows = openpyxl.load_workbook(f"{output_file}")[
            output_sheet_name
        ].max_row

        # append input.xlsx to output.xlsx
        workbook = openpyxl.load_workbook(f"{output_file}")
        worksheet = workbook[output_sheet_name]
        workbook2 = openpyxl.load_workbook(f"{input_file}")
        worksheet2 = workbook2[sheet_name]
        for row in worksheet2.iter_rows(min_row=2):
            # get all values from row
            values = [cell.value for cell in row]
            # append row to output.xlsx
            worksheet.append(values)
        workbook.save(f"{output_file}")
    else:
        original_max_rows = 0

    # copy original excel file called output.xlsx use shutil
    if not isExists:  # 如果BI檔不存在，建立一個新的BI檔
        shutil.copyfile(f"{input_file}", f"{output_file}")

    # load output.xlsx
    workbook = openpyxl.load_workbook(f"{output_file}")

    # 讀取sheet
    if not isExists:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[output_sheet_name]

    # 命名輸出的sheet名稱
    worksheet.title = output_sheet_name

    # get max row and column
    max_row = worksheet.max_row

    # max col is input.xlsx's
    max_column = openpyxl.load_workbook(f"{input_file}")[sheet_name].max_column

    # 執行excel的邏輯判斷式
    for row in tqdm(range(original_max_rows + 1, max_row + 1)):
        if row == 1 and not isExists:
            for i, col_name in enumerate(COL_NAMES):
                cell_col_name = worksheet.cell(row, max_column + i + 1)
                cell_col_name.value = COL_NAMES[i]

                # set background color
                cell_col_name.fill = openpyxl.styles.PatternFill(
                    fill_type="solid", fgColor="ED7D31"
                )

                # set width
                if not col_name:
                    # let col invisible
                    worksheet.column_dimensions[
                        get_column_letter(i + 106)
                    ].hidden = True
                else:
                    worksheet.column_dimensions[get_column_letter(i + 106)].width = 16
                    # set border line color
                    thin_border = Border(
                        left=Side(border_style=borders.BORDER_THIN, color="FF000000"),
                        right=Side(border_style=borders.BORDER_THIN, color="FF000000"),
                        top=Side(border_style=borders.BORDER_THIN, color="FF000000"),
                        bottom=Side(border_style=borders.BORDER_THIN, color="FF000000"),
                    )
                    cell_col_name.border = thin_border

        for column in range(1, max_column + 1):
            if column == max_column and row > 1:
                # DB
                cell = worksheet.cell(row, column + 1)
                cell.value = None

                # DC
                cell = worksheet.cell(row, column + 2)
                cell.value = f'=IF(CP{row}<-10,CP{row},IF(ISERROR(AVERAGE(CN{row}:CR{row})),"",AVERAGE(CN{row}:CR{row})))'

                # DD
                cell = worksheet.cell(row, column + 3)
                cell.number_format = "0%"
                cell.value = f'=IF(AC{row}<>"",AC{row}/100,"")'

                # DE
                cell = worksheet.cell(row, column + 4)
                cell.number_format = "0%"
                cell.value = f'=IF(AW{row}<>"",AW{row}/100,"")'

                # DF
                cell = worksheet.cell(row, column + 5)
                cell.number_format = "0%"
                cell.value = f'=IF(BQ{row}<>"",BQ{row}/100,"")'

                # DG
                cell = worksheet.cell(row, column + 6)
                cell.value = f"=MAX(DD{row},DE{row},DF{row})"

                # DH
                cell = worksheet.cell(row, column + 7)
                cell.value = f'=IF(DG{row}=DD{row},W{row},IF(DG{row}=DE{row},X{row},IF(DG{row}=DF{row},Y{row},"")))'

                # DI
                cell = worksheet.cell(row, column + 8)
                cell.value = f"=VLOOKUP(G{row},#REF!,2,0)"

                # DJ
                cell = worksheet.cell(row, column + 9)
                cell.value = f'=IF(DC{row}>-10,"",IF(ISERROR(DC{row}),"",CONCATENATE(INT(DC{row}/5)*5+5,"~",INT(DC{row}/5)*5)))'

                # DK
                cell = worksheet.cell(row, column + 10)
                cell.value = None

                # DL
                cell = worksheet.cell(row, column + 11)
                cell.value = f'=IF(AND(OR(N{row}="5G",N{row}="I5G"),O{row}="5GNSA"),"5G True User",IF(OR(N{row}="2G",N{row}="3G",N{row}="4G",N{row}="I4G"),"4G",IF(AND(OR(N{row}="5G",N{row}="I5G"),O{row}<>"5GNSA"),"5G非TU","")))'

                # DM
                cell = worksheet.cell(row, column + 12)
                cell.value = f'=COUNTIFS(AD{row}:AF{row},">-105",AD{row}:AF{row},"<0")+COUNTIFS(AX{row}:AZ{row},">-105",AX{row}:AZ{row},"<0")+COUNTIFS(BR{row}:BT{row},">-105",BR{row}:BT{row},"<0")'

                # DN
                cell = worksheet.cell(row, column + 13)
                cell.value = f"=ROUND(MAX(DD{row},DE{row},DF{row})*100/5,0)*0.05"

                # DO
                cell = worksheet.cell(row, column + 14)
                cell.value = f'=IF(DC{row}>-10,"",ROUND(DC{row}/5,0)*5)'

                # DP
                cell = worksheet.cell(row, column + 15)
                cell.value = f'=IF(R2="作業","障礙",\
IF(R{row}="障礙","障礙",\
IF(R{row}="抗爭","抗爭",\
IF(R{row}="40055重大障礙","40055重大障礙",\
IF(R{row}="非TWM問題的障礙","非TWM問題的障礙",\
IF(U{row}=35806,"非TWM問題的障礙",\
IF( OR(AND(AJ{row}<>"",AJ{row}>0,AJ{row}<0.7),\
       AND(AK{row}<>"",AK{row}>0,AK{row}<0.7),\
       AND(AL{row}<>"",AL{row}>0,AL{row}<0.7),\
       AND(AP{row}<>"",AP{row}>0,AP{row}<0.7),\
       AND(AQ{row}<>"",AQ{row}>0,AQ{row}<0.7),\
       AND(AR{row}<>"",AR{row}>0,AR{row}<0.7),\
       AND(AS{row}<>"",AS{row}>0,AS{row}<0.7),\
       AND(AT{row}<>"",AT{row}>0,AT{row}<0.7),\
       AND(AU{row}<>"",AU{row}>0,AU{row}<0.7)),"障礙",\
IF( OR(AND(BD{row}<>"",BD{row}>0,BD{row}<0.7),\
       AND(BE{row}<>"",BE{row}>0,BE{row}<0.7),\
       AND(BF{row}<>"",BF{row}>0,BF{row}<0.7),\
       AND(BJ{row}<>"",BJ{row}>0,BJ{row}<0.7),\
       AND(BK{row}<>"",BK{row}>0,BK{row}<0.7),\
       AND(BL{row}<>"",BL{row}>0,BL{row}<0.7),\
       AND(BM{row}<>"",BM{row}>0,BM{row}<0.7),\
       AND(BN{row}<>"",BN{row}>0,BN{row}<0.7),\
       AND(BO{row}<>"",BO{row}>0,BO{row}<0.7)),"障礙",\
IF( OR(AND(BX{row}<>"",BX{row}>0,BX{row}<0.7),\
       AND(BY{row}<>"",BY{row}>0,BY{row}<0.7),\
       AND(BZ{row}<>"",BZ{row}>0,BZ{row}<0.7),\
       AND(CD{row}<>"",CD{row}>0,CD{row}<0.7),\
       AND(CE{row}<>"",CE{row}>0,CE{row}<0.7),\
       AND(CF{row}<>"",CF{row}>0,CF{row}<0.7),\
       AND(CG{row}<>"",CG{row}>0,CG{row}<0.7),\
       AND(CH{row}<>"",CH{row}>0,CH{row}<0.7),\
       AND(CI{row}<>"",CI{row}>0,CI{row}<0.7)),"障礙",\
IF(OR(CJ{row}="住抗",CJ{row}="暫時移除設備"),"抗爭",\
IF(CJ{row}<>"","障礙",\
IF(DM{row}>2,"干擾",\
IF(Q{row}=6,"CC6",\
IF( OR(AND(DD{row}<>"",DD{row}>0.8),AND(DE{row}<>"",DE{row}>0.8),AND(DF{row}<>"",DF{row}>0.8)),"PRB>80",\
IF(AND(DC{row}>-106,DC{row}<-30),"RSRP優於-106",\
IF(DC{row}<=-106,"RSRP劣於-106",\
""))))))))))))))))'
                # DQ
                cell = worksheet.cell(row, column + 16)
                cell.value = f'=IF(ISERROR(SEARCH(">>檢查",AA{row})),"",MID(AA{row},SEARCH("PM分析:",AA{row})+5,SEARCH(">>檢查",AA{row})-SEARCH("PM分析:",AA{row})-5))'

                # DR
                cell = worksheet.cell(row, column + 17)
                cell.value = f'=IF(T{row}="因客訴地點人多，導致收訊擁擠","基站擁擠",\
IF(T{row}="因應特別活動調整相關參數導致","TTC",\
IF(OR(T{row}="基站障礙問題查測中",T{row}="基站問題待料中",T{row}="基站障礙問題已修復",T{row}="施工作業已恢復",T{row}="基站抗爭暫時關閉",T{row}="基站抗爭持續關閉中",T{row}="基站抗爭已復站",T{row}="基地台抗爭拆站",T{row}="基地台群體抗爭",T{row}="基站隱藏性障礙問題已修復"),"基站障礙",\
IF(OR(R{row}="作業",R{row}="障礙",R{row}="抗爭"),"基站障礙",\
IF(OR(T{row}="外在不明干擾影響，查測中",T{row}="干擾問題已排除",T{row}="外在不明干擾(大規模)影響",T{row}="干擾(大規模)問題已排除"),"干擾",\
IF(R{row}="干擾","干擾",\
""))))))'
        time.sleep(0.01)
    # save file
    workbook.save(f"{output_file}")

    # 設定樣式，比照原有的BI檔樣式
    if isExists:
        workbook = openpyxl.load_workbook(f"{output_file}")
        worksheet = workbook[output_sheet_name]
        for row in tqdm(
            worksheet.iter_rows(
                min_row=original_max_rows + 1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ),
            total=worksheet.max_row - original_max_rows,
        ):
            for cell in row:
                # set every cell style as its second row
                if not cell.row % 2:
                    cell._style = worksheet.cell(2, cell.column)._style
                else:
                    cell._style = worksheet.cell(3, cell.column)._style
            time.sleep(0.01)
        workbook.save(f"{output_file}")


if __name__ == "__main__":
    # 讀取ini檔
    configparser = ConfigParser()
    configparser.read("setting.ini", encoding="utf-8")
    section = "file_info"
    if configparser.has_section(section):
        file_info = dict(configparser.items(section))
    input_file = f"{file_info['input_file_path']}{file_info['input_file_name']}"
    output_file_name = f"{file_info['output_file_name']}"
    output_file_path = f"{file_info['output_file_path']}"
    output_sheet_name = f"{file_info['output_sheet_name']}"

    main(
        f"{input_file}",
        file_info["sheet_name"],
        f"{output_file_path}{output_file_name}",
        output_sheet_name,
    )
